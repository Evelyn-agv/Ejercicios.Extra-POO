import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    def excel():
        archivo = open("taller.txt","r",encoding="utf8")
        print(archivo)
        if "taller.txt" in os.listdir():
            print("El archivo se encuentra en el directorio")
        else:
            print("El archivo no se encuentra en el directorio")
        colocar = "./listado.xlsx"
        archivos = load_workbook(colocar)
        hoja = archivos.active
        estudiante = []
        for i, fila in enumerate(archivo):
            estudiante.append(fila)
            estudiante[i] = fila.replace("\n", "").split(';')
        archivo.close()
        for i, lista in enumerate(estudiante):
            estudiante[i][1] = lista[1].rstrip()
        for fila in estudiante:
            hoja.append(fila)
        columna(hoja, estudiante)
        archivos.save(colocar)
    def columna(hoja, datos):
        columnanueva = []
        for fila in datos:
            for i, x in enumerate(fila):
                if len(columnanueva) > i:
                    if len(x) > columnanueva[i]:
                        columnanueva[i] = len(x)
                else:
                    columnanueva += [len(x) + 5]

        for i, column_nu in enumerate(columnanueva):
            hoja.column_dimensions[get_column_letter(i + 1)].nu = column_nu

except ValueError:
    print("Error al trabajar con el archivo")

if __name__ == "__main__":
    excel()
import openpyxl
try:
    archivo=openpyxl.Workbook()
    hoja=archivo.active
    hoja.title="Estudiante"
    archivo.create_sheet("Notas")
    print(archivo.sheetnames)
except Exception as ex:
    print("Error al trabajar con el archivo")
    print(ex)
